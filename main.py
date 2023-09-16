import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename,name):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_unaligned_price = cell.value * 0.9
        corrected_price = sheet.cell(row,4)
        corrected_price.value = corrected_unaligned_price

    sheet.cell(1,4).value = 'Corrected Value'

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(f'{name}2.xlsx')

def main():
    name_file = input("File Name : ")
    filename = name_file+'.xlsx'
    process_workbook(filename,name_file)

if __name__ == "__main__":
    main()